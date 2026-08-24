import glob
import os
import pandas as pd
import openpyxl

# File Paths & Configuration
MASTER_DB_PATH = "VAPAHCS CPET DB.xlsx"
MASTER_SHEET_NAME = "baseline (new)"
BXB_FILES_DIRECTORY = "./CPET results (new)"  # Folder containing breath-by-breath Excel files


def extract_header_metadata(df_raw):
    """Scans the header section (rows 0-14) to extract patient demographics and test details."""
    key_mapping = {
        "ID1": "Study ID #",
        "Last Name": "Last Name",
        "First Name": "First Name",
        "Gender": "Gender",
        "Age": "Age at ETT",
        "Height (in)": "Height\n (in.)",
        "Weight (lbs)": "Weight \n(lbs.)",
        "Test date": "Date of ETT",
        "Reason for Stopping Test": "Angina (Ex or R) reason for stopping?",
        "Reason for Test": "Angina (Ex or R) reason for stopping?",
    }

    metadata = {}
    for r in range(min(15, len(df_raw))):
        for c in range(df_raw.shape[1]):
            val = df_raw.iloc[r, c]
            if pd.notna(val):
                key = str(val).strip()
                if key in key_mapping:
                    target_col = key_mapping[key]
                    if target_col in metadata:
                        continue  # Avoid overwriting primary reason for stopping

                    # Look right in the row for the associated value
                    for nc in range(c + 1, df_raw.shape[1]):
                        next_val = df_raw.iloc[r, nc]
                        if pd.notna(next_val):
                            next_str = str(next_val).strip()
                            if next_str in key_mapping:
                                break  # Stop if scanning into another metadata header label
                            if next_str != "":
                                metadata[target_col] = next_val
                                break
    return metadata


def process_cpet_file(file_path):
    """Parses a single CPET BxB Excel file and isolates peak exercise metrics."""
    xl = pd.ExcelFile(file_path)
    sheet_name = "Data" if "Data" in xl.sheet_names else xl.sheet_names[0]
    df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # 1. Extract Metadata Demographics
    meta = extract_header_metadata(df_raw)

    # Combine First/Last Name into "First Last" format
    first_name = str(meta.pop("First Name", "")).strip()
    last_name = str(meta.pop("Last Name", "")).strip()
    full_name = f"{first_name} {last_name}".strip()

    # Format Date
    date_val = meta.get("Date of ETT")
    formatted_date = (
        pd.to_datetime(date_val).strftime("%Y-%m-%d")
        if pd.notna(date_val)
        else None
    )

    # 2. Extract Breath-by-Breath Table Data (Header at Row 0, Col 9+)
    table_headers = list(df_raw.iloc[0, 9:].values)
    df_table = df_raw.iloc[3:, 9:].copy()
    df_table.columns = table_headers

    # 3. Locate Peak VO2 Row Safely
    vo2_col = "VO2/kg" if "VO2/kg" in df_table.columns else "VO2"
    df_table[vo2_col] = pd.to_numeric(df_table[vo2_col], errors="coerce")
    df_table = df_table.dropna(subset=[vo2_col]).reset_index(drop=True)

    max_idx = df_table[vo2_col].idxmax()
    max_row = df_table.iloc[max_idx].to_dict()

    # 4. Map Extracted Data
    extracted_data = {
        "Name": full_name if full_name else None,
        "Study ID #": meta.get("Study ID #"),
        "Date of ETT": formatted_date,
        "Gender": meta.get("Gender"),
        "Age at ETT": meta.get("Age at ETT"),
        "Height\n (in.)": meta.get("Height\n (in.)"),
        "Weight \n(lbs.)": meta.get("Weight \n(lbs.)"),
        "VO2 ml/Kg/min (max)": max_row.get("VO2/kg"),
        "VO2 in ml (max)": max_row.get("VO2"),
        "VCO2 (max)": max_row.get("VCO2"),
        "VE (max)": max_row.get("VE"),
        "RER max": max_row.get("RQ")
        if pd.notna(max_row.get("RQ"))
        else max_row.get("RER"),
        "Measured METs (Max VO2)": max_row.get("METS"),
        "VE/VC02 Slope": max_row.get("VE/VCO2"),
        "Angina (Ex or R) reason for stopping?": meta.get(
            "Angina (Ex or R) reason for stopping?"
        ),
    }

    return extracted_data


def normalize_val(val):
    """Normalizes string or date values for robust index lookup comparison."""
    if pd.isna(val) or val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    return str(val).strip().lower()


def update_master_db(records):
    """Locates existing patient/test date rows in Master DB and updates extracted metrics."""
    wb = openpyxl.load_workbook(MASTER_DB_PATH)
    ws = wb[MASTER_SHEET_NAME]

    # Map column headers to 1-based column index
    header_col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        col_name = ws.cell(row=1, column=col_idx).value
        if col_name is not None:
            header_col_map[str(col_name).strip()] = col_idx

    # Verify key identification columns exist
    id_col = header_col_map.get("Study ID #")
    name_col = header_col_map.get("Name")
    date_col = header_col_map.get("Date of ETT")

    if not date_col or (not id_col and not name_col):
        raise ValueError("Could not find required ID/Name and Date of ETT columns in sheet headers.")

    # Build row lookup map: (study_id or name, formatted_date) -> row_number
    row_index_map = {}
    for row_idx in range(2, ws.max_row + 1):
        d_val = normalize_val(ws.cell(row=row_idx, column=date_col).value)
        if not d_val:
            continue

        if id_col:
            id_val = normalize_val(ws.cell(row=row_idx, column=id_col).value)
            if id_val:
                row_index_map[(id_val, d_val)] = row_idx

        if name_col:
            name_val = normalize_val(ws.cell(row=row_idx, column=name_col).value)
            if name_val:
                row_index_map[(name_val, d_val)] = row_idx

    updated_count = 0
    not_found_count = 0

    for rec in records:
        rec_id = normalize_val(rec.get("Study ID #"))
        rec_name = normalize_val(rec.get("Name"))
        rec_date = normalize_val(rec.get("Date of ETT"))

        target_row = row_index_map.get((rec_id, rec_date)) or row_index_map.get((rec_name, rec_date))

        if target_row:
            # Update values in matching row
            for col_header, val in rec.items():
                if val is not None and col_header in header_col_map:
                    c_idx = header_col_map[col_header]
                    ws.cell(row=target_row, column=c_idx).value = val
            updated_count += 1
        else:
            print(f"Warning: No matching record found for ID: '{rec.get('Study ID #')}', Name: '{rec.get('Name')}', Date: '{rec.get('Date of ETT')}'")
            not_found_count += 1

    wb.save(MASTER_DB_PATH)
    print(f"Update Complete! Updated: {updated_count} record(s). Not found: {not_found_count} record(s).")


# Execution Pipeline
if __name__ == "__main__":
    bxb_files = glob.glob(
        os.path.join(BXB_FILES_DIRECTORY, "*BxB*.xlsx")
    ) + glob.glob(os.path.join(BXB_FILES_DIRECTORY, "*BxB*.xls"))

    records_to_update = []
    for file in bxb_files:
        if os.path.basename(file).startswith("~$"):  # Skip Excel temporary lock files
            continue
        try:
            record = process_cpet_file(file)
            records_to_update.append(record)
            print(f"Processed: {file}")
        except Exception as e:
            print(f"Error processing {file}: {e}")

    if records_to_update:
        update_master_db(records_to_update)
    else:
        print("No BxB files found to process.")