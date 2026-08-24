"""
match_ICD9_to_cpet.py

Fills the "history of / risk factor" columns in testcpetdb.xlsm using
diagnosis codes from Patients_All_ICD9_Codes.xlsx.
"""

import csv
import re
from copy import copy
from datetime import datetime, timedelta

import openpyxl
from openpyxl.utils import get_column_letter

# ============================== CONFIG ======================================

ICD9_FILE = "Patients_All_ICD9_Codes_Codes.xlsx"
ICD9_SHEET = None          # None = use the first sheet
ICD9_HEADER_ROW = 1

CPET_FILE = "cpet_filled_4.xlsx"
CPET_SHEET = None           # None = use the first sheet ("Sheet1" in the real file)
CPET_HEADER_ROW = 1

OUTPUT_FILE = "cpet_filled_new.xlsx"

# Two ICD9 dates within this many days of each other count as one soft
# match automatically. Anything wider goes to "Needs Review" instead of
# being silently accepted.
BIRTHDATE_TOLERANCE_DAYS = 1

# --- testcpetdb columns used for matching -----------------------------
CPET_NAME_COLUMN_CANDIDATES = ["Name", "Patient Name"]
CPET_TESTDATE_COLUMN_CANDIDATES = ["Date of ETT", "Test Date", "TestDate"]
CPET_AGE_COLUMN_CANDIDATES = ["Age at ETT", "Age"]

# --- Patients_All_ICD9_Codes.xlsx columns ---------------------------------
ICD9_FIRSTNAME_COL = "PatientFirstName"
ICD9_LASTNAME_COL = "PatientLastName"
ICD9_BIRTHDATE_COL = "BirthDateTime"
ICD9_CODE_COL = "ICD9Code"
ICD9_DESC_COL = "ICD9Description"
ICD9_DATE_COL = "Earliest_ICD9Code_Visit_Time"

MAPPING_CSV = "ICD9_category_mapping.csv"

# Target columns in testcpetdb, keyed by clinical category.
CATEGORY_FIELDS = {
    "HTN":          {"flag": "H/o__hypertension (HTN)",                 "date": "HTN_date of onset"},
    "CHF":          {"flag": "h/o__Congestive heart failure (CHF)",     "date": "CHF_date of onset"},
    "CM":           {"flag": "h/o__Cardiomyopathy (CM)",                "date": "CM_date of onset"},
    "MI":           {"flag": "h/o__Myocardial Infarction (MI)",         "date": "MI_date of onset"},
    "Stroke":       {"flag": "h/o_stroke",                              "date": "Stroke_date of onset"},
    "CAD":          {"flag": "h/o__Coronary artery disease (CAD)",      "date": "Cad_date of onset"},
    "PAD":          {"flag": "h/o__peripheral artereial disease (PAD)", "date": "PAD_date of onset"},
    "Diabetes":     {"flag": "h/o__diabetes_id",                        "date": "Diabetes_date of onset"},
    "CKD":          {"flag": "h/o_Chronic kideny disease (CKD)",        "date": "CKD onset date"},
    "Dyslipidemia": {"flag": "risk_factors_dyslipdemia",                "date": "dyslipidemia_date of onset"},
    "Smoke":        {"flag": "risk_factors__smoke_id",                  "date": "Smoke_date of onset"},
    "Alcohol":      {"flag": "risk_factors__alcohol",                   "date": "Alcohol_date of onset"},
    "Drug":         {"flag": "Risk factor_drug",                        "date": "Drug_date of onset"},
}

SMOKE_CURRENT_PRIOR_HEADER = "risk_factors__smoke_current/prior"
SMOKE_PACKYEARS_HEADER = "risk_factors__smoke_packsyears"

# =============================== END CONFIG =================================


def normalize_name(first, last, middle_ok=True):
    """Lowercase, strip punctuation/whitespace. Returns (first_token, last_token)."""
    def clean(s):
        s = (s or "").strip().lower()
        s = re.sub(r"[.,'\-]", "", s)
        s = re.sub(r"\s+", " ", s)
        return s
    return clean(first).split(" ")[0] if clean(first) else "", clean(last).split(" ")[-1] if clean(last) else ""


def parse_combined_name(raw):
    """Best-effort split of a single 'Name' column into (first, last)."""
    if not raw:
        return "", ""
    raw = str(raw).strip()
    if "," in raw:
        last, rest = raw.split(",", 1)
        tokens = rest.strip().split()
        first = tokens[0] if tokens else ""
        return first, last.strip()
    tokens = raw.split()
    if len(tokens) == 1:
        return tokens[0], ""
    return tokens[0], tokens[-1]


def find_header_index(header_row_values, candidates):
    """Return 0-based column index whose header contains any candidate string (case-insensitive)."""
    for idx, val in enumerate(header_row_values):
        if val is None:
            continue
        text = str(val).strip().lower()
        for c in candidates:
            if c.lower() in text:
                return idx
    return None


def find_exact_header_index(header_row_values, exact_text):
    """Return 0-based column index whose header exactly matches exact_text."""
    target = exact_text.strip().lower()
    for idx, val in enumerate(header_row_values):
        if val is None:
            continue
        if str(val).strip().lower() == target:
            return idx
    return None


def to_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if hasattr(value, "year"):
        return datetime(value.year, value.month, value.day)
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            continue
    return None


def load_mapping(path):
    mapping = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            prefix = row["Prefix"].strip()
            category = row["Category"].strip()
            if prefix and category:
                mapping.append((prefix, category))
    mapping.sort(key=lambda x: -len(x[0]))
    return mapping


def code_to_category(code, mapping):
    if not code:
        return None
    code = str(code).strip().upper()
    for prefix, category in mapping:
        if code.startswith(prefix.upper()):
            return category
    return None


def load_ICD9_patients(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[ICD9_SHEET] if ICD9_SHEET else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[ICD9_HEADER_ROW - 1]]
    col = {name: header.index(name) for name in
           [ICD9_FIRSTNAME_COL, ICD9_LASTNAME_COL, ICD9_BIRTHDATE_COL,
            ICD9_CODE_COL, ICD9_DESC_COL, ICD9_DATE_COL] if name in header}

    missing = [n for n in [ICD9_FIRSTNAME_COL, ICD9_LASTNAME_COL, ICD9_BIRTHDATE_COL,
                            ICD9_CODE_COL, ICD9_DESC_COL, ICD9_DATE_COL] if n not in col]
    if missing:
        raise ValueError(f"ICD9 file is missing expected columns: {missing}")

    patients = {}
    for r in rows[ICD9_HEADER_ROW:]:
        if r is None or all(v is None for v in r):
            continue
        first = r[col[ICD9_FIRSTNAME_COL]]
        last = r[col[ICD9_LASTNAME_COL]]
        bdate = to_date(r[col[ICD9_BIRTHDATE_COL]])
        code = r[col[ICD9_CODE_COL]]
        desc = r[col[ICD9_DESC_COL]]
        vdate = to_date(r[col[ICD9_DATE_COL]])
        if not first and not last:
            continue
        nfirst, nlast = normalize_name(first, last)
        key = (nfirst, nlast, bdate)
        patients.setdefault(key, []).append({"code": code, "desc": desc, "date": vdate})
    return patients


def build_name_index(patients):
    idx = {}
    for (nfirst, nlast, bdate) in patients:
        idx.setdefault((nfirst, nlast), []).append(bdate)
    return idx


def copy_row_style_only(ws, src_row, dst_row, ncols):
    """Copies cell formatting/styling without copying data values."""
    for c in range(1, ncols + 1):
        src_cell = ws.cell(row=src_row, column=c)
        dst_cell = ws.cell(row=dst_row, column=c)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)


def fill_indication_cell(ws, target_row, cat, item, field_cols, smoke_cp_idx):
    """Helper function to set indication flag, date, and smoking status for a row."""
    fcol = field_cols[cat]["flag"]
    dcol = field_cols[cat]["date"]
    
    icd_code_str = str(item["code"] or "").strip()
    ws.cell(row=target_row, column=fcol + 1).value = f"Y - {icd_code_str}" if icd_code_str else "Y"
    
    if item["date"]:
        ws.cell(row=target_row, column=dcol + 1).value = item["date"]
    if cat == "Smoke" and smoke_cp_idx is not None:
        code = icd_code_str.upper()
        if code.startswith("Z87.891"):
            ws.cell(row=target_row, column=smoke_cp_idx + 1).value = "Prior"
        elif code.startswith("F17"):
            ws.cell(row=target_row, column=smoke_cp_idx + 1).value = "Current"


def main():
    print("Loading ICD9 data ...")
    mapping = load_mapping(MAPPING_CSV)
    ICD9_patients = load_ICD9_patients(ICD9_FILE)
    name_index = build_name_index(ICD9_patients)
    print(f"  {len(ICD9_patients)} unique (name, birthdate) patients loaded from ICD9 file.")

    print("Loading testcpetdb workbook ...")
    wb = openpyxl.load_workbook(CPET_FILE)
    ws = wb[CPET_SHEET] if CPET_SHEET else wb.worksheets[0]

    header_vals = [c.value for c in ws[CPET_HEADER_ROW]]
    header_vals_clean = [str(h).strip() if h is not None else "" for h in header_vals]

    name_col = find_header_index(header_vals_clean, CPET_NAME_COLUMN_CANDIDATES)
    testdate_col = find_header_index(header_vals_clean, CPET_TESTDATE_COLUMN_CANDIDATES)
    age_col = find_header_index(header_vals_clean, CPET_AGE_COLUMN_CANDIDATES)

    if name_col is None or testdate_col is None or age_col is None:
        raise ValueError("Could not find Name / Test Date / Age columns in testcpetdb.")

    field_cols = {}
    missing_headers = []
    for category, fields in CATEGORY_FIELDS.items():
        flag_idx = find_exact_header_index(header_vals_clean, fields["flag"])
        date_idx = find_exact_header_index(header_vals_clean, fields["date"])
        if flag_idx is None:
            missing_headers.append(fields["flag"])
        if date_idx is None:
            missing_headers.append(fields["date"])
        field_cols[category] = {"flag": flag_idx, "date": date_idx}
    smoke_cp_idx = find_exact_header_index(header_vals_clean, SMOKE_CURRENT_PRIOR_HEADER)
    if smoke_cp_idx is None:
        missing_headers.append(SMOKE_CURRENT_PRIOR_HEADER)

    if missing_headers:
        raise ValueError(f"Expected column headers were not found: {missing_headers}")

    ncols = len(header_vals_clean)
    first_data_row = CPET_HEADER_ROW + 1
    last_data_row = ws.max_row

    # Pass 1: build consecutive patient blocks
    blocks = []
    prev_key = None
    for r in range(first_data_row, last_data_row + 1):
        raw_name = ws.cell(row=r, column=name_col + 1).value
        first, last = parse_combined_name(raw_name)
        nfirst, nlast = normalize_name(first, last)
        key = (nfirst, nlast)
        if key == ("", ""):
            prev_key = None
            continue
        if key == prev_key:
            blocks[-1]["end_row"] = r
        else:
            blocks.append({"start_row": r, "end_row": r, "first": first, "last": last,
                            "nfirst": nfirst, "nlast": nlast})
        prev_key = key

    print(f"Found {len(blocks)} distinct patient blocks across {last_data_row - first_data_row + 1} rows.")

    review_rows = []
    matched_count = 0
    inserted_rows = 0

    # Pass 2: process bottom-to-top
    for block in reversed(blocks):
        start_r = block["start_row"]
        end_r = block["end_row"]
        available_rows = (end_r - start_r) + 1

        test_date = to_date(ws.cell(row=start_r, column=testdate_col + 1).value)
        age_val = ws.cell(row=start_r, column=age_col + 1).value
        try:
            age = float(age_val)
        except (TypeError, ValueError):
            age = None

        approx_birthdate = None
        if test_date and age is not None:
            approx_birthdate = test_date - timedelta(days=age * 365.25)

        candidates = name_index.get((block["nfirst"], block["nlast"]))

        match_key = None
        status = "no_match"
        day_gap = None

        if candidates and approx_birthdate:
            best = None
            best_gap = None
            for bdate in candidates:
                if bdate is None:
                    continue
                gap = abs((bdate - approx_birthdate).days)
                if best_gap is None or gap < best_gap:
                    best_gap = gap
                    best = bdate
            if best is not None:
                day_gap = best_gap
                if best_gap <= BIRTHDATE_TOLERANCE_DAYS:
                    match_key = (block["nfirst"], block["nlast"], best)
                    status = "matched"
                else:
                    status = "soft_no_match"
        elif candidates and not approx_birthdate:
            status = "no_birthdate"

        if status != "matched":
            review_rows.append({
                "row": start_r, "first": block["first"], "last": block["last"],
                "test_date": test_date, "age": age, "reason": status,
                "day_gap": day_gap,
            })
            continue

        matched_count += 1
        icd_rows = ICD9_patients[match_key]

        by_category = {}
        for item in icd_rows:
            cat = code_to_category(item["code"], mapping)
            if cat:
                by_category.setdefault(cat, []).append(item)

        # Sort items chronologically within each category
        for cat in by_category:
            by_category[cat].sort(key=lambda x: (x["date"] is None, x["date"]))

        # Fill existing patient rows first
        for cat, items in by_category.items():
            for idx, item in enumerate(items):
                if idx < available_rows:
                    target_row = start_r + idx
                    fill_indication_cell(ws, target_row, cat, item, field_cols, smoke_cp_idx)

        # Calculate if new rows are required beyond the block's current size
        max_items_needed = max([len(items) for items in by_category.values()], default=0)
        extra_needed = max_items_needed - available_rows

        # Insert secondary rows only if existing rows ran out
        if extra_needed > 0:
            insert_at = end_r + 1
            for i in range(extra_needed):
                idx = available_rows + i
                ws.insert_rows(insert_at)
                copy_row_style_only(ws, start_r, insert_at, ncols)
                
                for cat, items in by_category.items():
                    if len(items) > idx:
                        item = items[idx]
                        fill_indication_cell(ws, insert_at, cat, item, field_cols, smoke_cp_idx)

                inserted_rows += 1
                insert_at += 1

    # Needs Review sheet
    if "Needs Review" not in wb.sheetnames:
        review_ws = wb.create_sheet("Needs Review")
        review_ws.append(["cpet Row", "First", "Last", "Test Date", "Age", "Reason", "Birthdate Day Gap"])
    else:
        review_ws = wb["Needs Review"]
        review_ws.append([])  # Adds a blank line break before appending new rows

    for rr in sorted(review_rows, key=lambda x: x["row"]):
        review_ws.append([rr["row"], rr["first"], rr["last"],
                          rr["test_date"].date() if rr["test_date"] else None,
                          rr["age"], rr["reason"], rr["day_gap"]])

    wb.save(OUTPUT_FILE)

    print("---------------------------------------------------------------")
    print(f"Matched and filled: {matched_count} patient blocks")
    print(f"Extra rows inserted for multiple codes: {inserted_rows}")
    print(f"Sent to 'Needs Review': {len(review_rows)} patient blocks")
    print(f"Saved: {OUTPUT_FILE}")
    print("---------------------------------------------------------------")


if __name__ == "__main__":
    main()